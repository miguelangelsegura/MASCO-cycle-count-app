"""
Masco Canada — Cycle Count Web App v2
Multi-user, persistent sessions with dashboard
"""

from flask import Flask, render_template, request, jsonify, send_file
import csv, io, json, os, sqlite3, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cyclecount.db")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_num TEXT NOT NULL,
            description TEXT,
            created_at TEXT,
            status TEXT DEFAULT 'active',
            items_json TEXT,
            counts_json TEXT DEFAULT '{}'
        )
    """)
    conn.commit()
    conn.close()

init_db()

def parse_jde_csv(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    cycle_num, description = "UNKNOWN", "UNKNOWN"
    for r in all_rows[:7]:
        if r and r[0].strip() == "Cycle Count Number" and len(r) > 2:
            cycle_num = r[2].strip()
        if r and r[0].strip() == "Description" and len(r) > 2:
            description = r[2].strip()
    items = []
    for r in all_rows[7:]:
        if not r or not r[0].strip():
            continue
        try:
            qoh = int(float(r[3].strip())) if len(r) > 3 and r[3].strip() else 0
        except:
            qoh = 0
        items.append({
            "item":     r[0].strip(),
            "qoh":      qoh,
            "desc":     r[7].strip() if len(r) > 7 else "",
            "location": r[14].strip() if len(r) > 14 else "",
            "abc":      r[10].strip() if len(r) > 10 else "",
            "branch":   r[12].strip().strip() if len(r) > 12 else "MC08",
        })
    return cycle_num, description, items

def build_export(batch_row):
    items  = json.loads(batch_row["items_json"])
    counts = json.loads(batch_row["counts_json"] or "{}")
    batch  = batch_row["batch_num"]
    desc   = batch_row["description"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JDE_PASTE"
    NAVY="1B3A6B"; TEAL="1E7B74"; WHITE="FFFFFF"
    GREEN="1A7A3C"; LTGREEN="E8F5EE"; GRAY="4A4A4A"; LTGRAY="F2F2F2"; MID="CCCCCC"
    def s(c, w="thin"):
        x = Side(style=w, color=c)
        return Border(left=x, right=x, top=x, bottom=x)
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"Batch {batch} — {desc} — Copy col A → Paste into JDE Quantity"
    t.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    t.fill = PatternFill("solid", fgColor=TEAL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for c, (h, w) in enumerate(zip(["QUANTITY","Item Number","Description"],[14,20,40]),1):
        cell = ws.cell(2, c, h)
        cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = s(MID)
        ws.column_dimensions[get_column_letter(c)].width = w
    for i, item in enumerate(items, 3):
        key = str(i - 3)
        c_data = counts.get(key, {})
        upb = c_data.get("upb", 0); boxes = c_data.get("boxes", 0)
        total = upb * boxes if upb and boxes else 0
        bg = LTGRAY if i % 2 == 0 else WHITE
        qty = ws.cell(i, 1, total)
        qty.font = Font(bold=True, size=12, name="Calibri", color=GREEN if total > 0 else GRAY)
        qty.fill = PatternFill("solid", fgColor=LTGREEN if total > 0 else bg)
        qty.alignment = Alignment(horizontal="center", vertical="center")
        qty.border = s(TEAL if total > 0 else MID, "medium" if total > 0 else "thin")
        for c, val in enumerate([item["item"], item["desc"]], 2):
            cell = ws.cell(i, c, val)
            cell.font = Font(size=10, name="Calibri", bold=(c==2), color=NAVY if c==2 else GRAY)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = s(MID)
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/batches")
def get_batches():
    conn = get_db()
    rows = conn.execute("SELECT id, batch_num, description, created_at, status, items_json, counts_json FROM batches ORDER BY created_at DESC").fetchall()
    conn.close()
    result = []
    for r in rows:
        items  = json.loads(r["items_json"] or "[]")
        counts = json.loads(r["counts_json"] or "{}")
        done   = sum(1 for v in counts.values() if v.get("upb") and v.get("boxes"))
        result.append({"id": r["id"], "batch_num": r["batch_num"], "description": r["description"],
                        "created_at": r["created_at"], "status": r["status"],
                        "total_items": len(items), "done_items": done})
    return jsonify(result)

@app.route("/api/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f: return jsonify({"error": "No file"}), 400
    try:
        batch, description, items = parse_jde_csv(f.read())
        conn = get_db()
        existing = conn.execute("SELECT id FROM batches WHERE batch_num=?", (batch,)).fetchone()
        if existing:
            conn.execute("UPDATE batches SET items_json=?, description=?, status='active' WHERE batch_num=?",
                         (json.dumps(items), description, batch))
            conn.commit(); batch_id = existing["id"]
        else:
            cursor = conn.execute("INSERT INTO batches (batch_num, description, created_at, items_json, counts_json) VALUES (?,?,?,?,?)",
                (batch, description, datetime.now().strftime("%Y-%m-%d %H:%M"), json.dumps(items), "{}"))
            conn.commit(); batch_id = cursor.lastrowid
        conn.close()
        return jsonify({"id": batch_id, "batch": batch, "description": description, "count": len(items)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/batch/<int:batch_id>")
def get_batch(batch_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM batches WHERE id=?", (batch_id,)).fetchone()
    conn.close()
    if not row: return jsonify({"error": "Not found"}), 404
    items  = json.loads(row["items_json"] or "[]")
    counts = json.loads(row["counts_json"] or "{}")
    done   = sum(1 for v in counts.values() if v.get("upb") and v.get("boxes"))
    return jsonify({"id": row["id"], "batch_num": row["batch_num"], "description": row["description"],
                    "status": row["status"], "items": items, "counts": counts,
                    "done": done, "total": len(items)})

@app.route("/api/count", methods=["POST"])
def save_count():
    data  = request.json
    bid   = data.get("batch_id")
    idx   = str(data.get("idx"))
    upb   = int(data.get("upb", 0))
    boxes = int(data.get("boxes", 0))
    conn  = get_db()
    row   = conn.execute("SELECT counts_json, items_json FROM batches WHERE id=?", (bid,)).fetchone()
    if not row: conn.close(); return jsonify({"error": "Not found"}), 404
    counts = json.loads(row["counts_json"] or "{}")
    counts[idx] = {"upb": upb, "boxes": boxes}
    total_items = len(json.loads(row["items_json"] or "[]"))
    done = sum(1 for v in counts.values() if v.get("upb") and v.get("boxes"))
    conn.execute("UPDATE batches SET counts_json=? WHERE id=?", (json.dumps(counts), bid))
    conn.commit(); conn.close()
    return jsonify({"total": upb*boxes, "done": done, "total_items": total_items})

@app.route("/api/batch/<int:batch_id>/complete", methods=["POST"])
def complete_batch(batch_id):
    conn = get_db()
    conn.execute("UPDATE batches SET status='completed' WHERE id=?", (batch_id,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/batch/<int:batch_id>", methods=["DELETE"])
def delete_batch(batch_id):
    conn = get_db()
    conn.execute("DELETE FROM batches WHERE id=?", (batch_id,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/export/<int:batch_id>")
def export(batch_id):
    conn = get_db()
    row  = conn.execute("SELECT * FROM batches WHERE id=?", (batch_id,)).fetchone()
    conn.close()
    if not row: return "Not found", 404
    buf = build_export(row)
    return send_file(buf, as_attachment=True,
                     download_name=f"CycleCount_{row['batch_num']}_JDE_Upload.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    import socket
    try: local_ip = socket.gethostbyname(socket.gethostname())
    except: local_ip = "your-computer-IP"
    print(f"\n{'='*55}\n  MASCO CANADA — Cycle Count App v2\n{'='*55}")
    print(f"  Don's computer : http://localhost:5000")
    print(f"  Counters (iPad): http://{local_ip}:5000\n{'='*55}\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
